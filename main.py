from fastapi import FastAPI
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, FileResponse
from pydantic import BaseModel
from typing import List, Optional
import uvicorn, os, math, re, PyPDF2, sqlite3, json
from fpdf import FPDF
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

app = FastAPI()
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_methods=["*"], allow_headers=["*"])

PDF_FOLDER = os.path.join(os.getcwd(), "tds")
DB_FILE = "molty.db"

# --- DATABASE INIT ---
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS projects
                 (id INTEGER PRIMARY KEY AUTOINCREMENT,
                  client TEXT,
                  date TEXT,
                  metal TEXT,
                  total_weight REAL,
                  total_cost REAL,
                  data TEXT)''')
    conn.commit()
    conn.close()

init_db() # Pokreni pri startu

# --- DATA MOCK ---
CLIENTS_DB = ["METALFER STEEL MILL", "HBIS GROUP", "ZIJIN BOR COPPER", "US STEEL KOSICE", "ARCELLOR MITTAL"]
METALS_DB = {
    "Celik (Low C)": 1510, "Sivi Liv": 1200, "Nodularni Liv": 1150,
    "Bakar": 1085, "Mesing": 930, "Bronza": 950, "Aluminijum": 660
}

def get_mats():
    mats = [{"name": "STEEL SHELL (S235)", "density": 7850, "lambda_val": 50.0, "price": 1000},
            {"name": "AIR GAP", "density": 1, "lambda_val": 0.05, "price": 0}]
    if os.path.exists(PDF_FOLDER):
        for f in os.listdir(PDF_FOLDER):
            if f.lower().endswith(".pdf"):
                try:
                    r = PyPDF2.PdfReader(os.path.join(PDF_FOLDER, f))
                    txt = r.pages[0].extract_text() or ""
                    dm = re.search(r"(\d+[.,]?\d*)\s*(kg/m3|g/cm3)", txt, re.IGNORECASE)
                    den = 2300
                    if dm:
                        val = float(dm.group(1).replace(",", "."))
                        den = val * 1000 if val < 10 else val
                    mats.append({"name": f[:-4].upper(), "density": int(den), "lambda_val": 1.5, "price": 800})
                except: pass
    return sorted(mats, key=lambda x: x["name"])

# --- MODELS ---
class Layer(BaseModel):
    material: str; thickness: float; lambda_val: float; density: float; price: float
class SimReq(BaseModel):
    metal: str; target_temp: float; ambient_temp: float; layers: List[Layer]; geometry: dict
    client: Optional[str] = "" # Dodato za bazu
class ExpReq(BaseModel):
    bom: List[dict]; total_weight: float; total_cost: float; shell_temp: float; heat_flux: float; client: str

# --- API ROUTES ---

@app.get("/api/init")
def init_data():
    return {"materials": get_mats(), "metals": METALS_DB, "clients": CLIENTS_DB}

# NOVO: CUVANJE PROJEKTA
@app.post("/api/db/save")
def save_project(r: SimReq):
    # Racunamo basic stats za pregled
    w = sum([l.thickness * l.density for l in r.layers]) # Aproksimacija za brz prikaz
    cost = sum([l.price for l in r.layers])
    
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("INSERT INTO projects (client, date, metal, total_weight, total_cost, data) VALUES (?, ?, ?, ?, ?, ?)",
              (r.client, datetime.now().strftime("%Y-%m-%d %H:%M"), r.metal, w, cost, json.dumps(r.dict())))
    conn.commit()
    last_id = c.lastrowid
    conn.close()
    return {"status": "saved", "id": last_id}

# NOVO: LISTA PROJEKATA
@app.get("/api/db/list")
def list_projects():
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT id, client, date, metal FROM projects ORDER BY id DESC")
    rows = c.fetchall()
    conn.close()
    return rows

# NOVO: UCITAVANJE
@app.get("/api/db/load/{pid}")
def load_project(pid: int):
    conn = sqlite3.connect(DB_FILE)
    conn.row_factory = sqlite3.Row
    c = conn.cursor()
    c.execute("SELECT data FROM projects WHERE id=?", (pid,))
    row = c.fetchone()
    conn.close()
    return json.loads(row['data']) if row else {}

@app.post("/api/simulate")
def calc(r: SimReq):
    t_hot = r.target_temp; t_amb = r.ambient_temp
    temps = [t_hot]; curr = t_hot
    for _ in range(3):
        total_r = 0.12
        r_vals = []
        for i, l in enumerate(r.layers):
            t_mean = (temps[i] if i < len(temps) else t_hot) * 0.9 
            lam = (l.lambda_val or 1.0) * (1 + (t_mean/2000)*0.1)
            res = (l.thickness/1000.0) / lam
            r_vals.append(res); total_r += res
        flux = (t_hot - t_amb) / total_r
        curr = t_hot; temps = [curr]
        for rv in r_vals: curr -= flux * rv; temps.append(curr)

    bom = []; acc_th = 0; cid = r.geometry['dim2']; tw = 0; tc = 0
    liquidus = METALS_DB.get(r.metal, 0)
    freeze_depth = -1
    
    for i, l in enumerate(r.layers):
        if temps[i] >= liquidus and temps[i+1] <= liquidus:
            ratio = (temps[i] - liquidus) / (temps[i] - temps[i+1])
            freeze_depth = acc_th + (l.thickness * ratio)
        acc_th += l.thickness

        w = 0
        if r.geometry['type'] == 'cylinder':
            od = cid + 2*l.thickness
            w = math.pi * r.geometry['dim1'] * ((od/1000/2)**2 - (cid/1000/2)**2) * l.density
            bom.append({"name":l.material, "th":l.thickness, "temp":round(temps[i+1]), "id":round(cid), "od":round(od), "w":round(w,1), "cost":round(w/1000*l.price,1)})
            cid = od
        else:
            w = r.geometry['dim1'] * (l.thickness/1000.0) * l.density
            bom.append({"name":l.material, "th":l.thickness, "temp":round(temps[i+1]), "id":"-", "od":"-", "w":round(w,1), "cost":round(w/1000*l.price,1)})
        tw += w; tc += w/1000*l.price

    status = "SAFE"
    if freeze_depth < 0: status = "CRITICAL (Proboj)"
    elif freeze_depth > (acc_th - bom[-1]['th']): status = "WARNING (Plašt)"

    return {
        "shell_temp": round(temps[-1], 1), "heat_flux": round(flux, 1),
        "profile": [{"pos": 0, "temp": t_hot}] + [{"pos": sum(x['th'] for x in bom[:i+1]), "temp": t} for i,t in enumerate(temps[1:])],
        "bom": bom, "total_weight": round(tw, 1), "total_cost": round(tc, 1),
        "liquidus": liquidus, "freeze_depth": round(freeze_depth, 1) if freeze_depth > 0 else "N/A", "safety": status,
        "req_shell": round(cid) if r.geometry['type']=='cylinder' else "-"
    }

# --- EXPORT RUTES (UNCHANGED) ---
@app.post("/api/export-excel")
def excel(d: ExpReq):
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Specifikacija"
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    ws['A1'] = "PROJEKAT:"; ws['B1'] = d.client; ws['A2'] = "DATUM:"; ws['B2'] = datetime.now().strftime("%Y-%m-%d")
    headers = ["POZICIJA", "MATERIJAL", "DEBLJINA (mm)", "TEMP. SPOJA (°C)", "KOLIČINA (kg)", "CENA (€/t)"]
    ws.append([]); ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num); cell.value = header; cell.font = header_font; cell.fill = header_fill; cell.alignment = Alignment(horizontal='center'); cell.border = border
    for i, r in enumerate(d.bom, 1):
        row = [i, r['name'], r['th'], r['temp'], r['w'], r['cost'] / (r['w']/1000) if r['w'] > 0 else 0]
        ws.append(row)
        for col in range(1, 7): 
            ws.cell(row=4+i, column=col).border = border
            if col == 6: ws.cell(row=4+i, column=col).fill = PatternFill(start_color='FFFFCC', fill_type='solid')
    ws.append([]); ws.append(["UKUPNO:", "", "", "", d.total_weight, f"=SUM(F5:F{4+len(d.bom)})"])
    wb.save("spec.xlsx"); return FileResponse("spec.xlsx", filename=f"Specifikacija_{d.client}.xlsx")

@app.post("/api/export-pdf")
def pdf(d: ExpReq):
    pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial","B",16); pdf.cell(0,10,f"PROJECT: {d.client}",0,1,"L")
    pdf.set_font("Arial","",10); pdf.cell(0,10,f"Date: {datetime.now().strftime('%Y-%m-%d')}",0,1,"L"); pdf.line(10,25,200,25); pdf.ln(10)
    pdf.set_font("Arial","B",10); pdf.set_fill_color(230,230,230)
    pdf.cell(80,8,"Material",1,0,"L",1); pdf.cell(20,8,"Thk",1,0,"C",1); pdf.cell(30,8,"Temp",1,0,"C",1); pdf.cell(30,8,"Kg",1,0,"R",1); pdf.cell(30,8,"Eur",1,1,"R",1)
    pdf.set_font("Arial","",9)
    for r in d.bom:
        pdf.cell(80,7,r['name'][:35].encode('latin-1','replace').decode('latin-1'),1); pdf.cell(20,7,str(r['th']),1,0,"C"); pdf.cell(30,7,str(r['temp']),1,0,"C"); pdf.cell(30,7,str(r['w']),1,0,"R"); pdf.cell(30,7,str(r['cost']),1,1,"R")
    pdf.ln(5); pdf.set_font("Arial","B",11); pdf.cell(0,10,f"TOTAL WEIGHT: {d.total_weight} kg",0,1); pdf.output("r.pdf"); return FileResponse("r.pdf", filename=f"Report_{d.client}.pdf")

@app.get("/", response_class=HTMLResponse)
def root(): 
    if os.path.exists("dashboard.html"): return open("dashboard.html", encoding="utf-8").read()
    return "Error"

if __name__ == "__main__": uvicorn.run(app, host="0.0.0.0", port=8000)
